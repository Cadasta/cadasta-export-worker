import os
import tempfile

from openpyxl import Workbook
from shapely.geometry import shape

from ..celery import app
from ..settings import PLATFORM_URL, QUEUE
from ..utils.api import fetch_data, upload_file, ZipStreamQueue
from ..utils import data as data_utils


@app.task(name='{}.project.xls'.format(QUEUE), bind=True)
def export_xls(self, org_slug, project_slug, api_key, bundle_url, out_dir):
    base_url = '{base}/api/v1/organizations/{org}/projects/{proj}'
    base_url = base_url.format(
        base=PLATFORM_URL, org=org_slug, proj=project_slug)

    filename = '{}.xlsx'.format(project_slug)
    prefix = os.path.join(org_slug, project_slug, self.request.id)
    tmp_prefix = '{}_'.format(prefix.replace('/', '_'))
    with tempfile.TemporaryDirectory(prefix=tmp_prefix) as tmpdir:
        wb = Workbook(write_only=True)

        write_locations_sheet(wb, base_url, api_key)
        write_parties_sheet(wb, base_url, api_key)
        write_relationships_sheet(wb, base_url, api_key)

        path = os.path.join(tmpdir, filename)
        wb.save(path)
        bucket, key = upload_file(os.path.join(prefix, filename), path)
        with ZipStreamQueue(bundle_url) as q:
            src = 's3://{bucket}/{key}'.format(bucket=bucket, key=key)
            q.insert(data_utils.get_zipstream_payload(src, out_dir))


def write_locations_sheet(wb, base_url, api_key):
    url = '{base}/spatial/'.format(base=base_url)

    headers = [
        ('id', 'properties.id'),
        ('geometry.ewkt', 'geometry'),
        ('type', 'properties.type'),
        ('area', 'properties.area'),
    ]
    header_len = len(headers)

    data = []
    for resp in fetch_data(api_key, url, array_response=False):
        for obj in resp['features']:
            headers = data_utils.append_missing_headers(
                obj, headers, attrs='properties.attributes')
            obj['geometry'] = shape(obj['geometry']).wkt
            data.append(obj)
    headers = data_utils.order_headers(headers, header_len)
    headers = data_utils.normalize_headers(headers)

    sheet = wb.create_sheet('locations')
    sheet.append([header[0].split('.')[-1] for header in headers])

    for obj in data:
        sheet.append([data_utils.get_attr(obj, attr[1]) for attr in headers])


def write_parties_sheet(wb, base_url, api_key):
    url = '{base}/parties/'.format(base=base_url)

    headers = ['id', 'name', 'type']
    header_len = len(headers)

    data = []
    for obj in fetch_data(api_key, url):
        headers = data_utils.append_missing_headers(obj, headers)
        data.append(obj)
    headers = data_utils.order_headers(headers, header_len)
    headers = data_utils.normalize_headers(headers)

    sheet = wb.create_sheet('parties')
    sheet.append([header[0].split('.')[-1] for header in headers])

    for obj in data:
        sheet.append([data_utils.get_attr(obj, attr[1]) for attr in headers])


def write_relationships_sheet(wb, base_url, api_key):
    url = '{base}/relationships/tenure/'.format(base=base_url)

    headers = [
        ('party_id', 'party'),
        ('spatial_unit_id', 'spatial_unit'),
        'tenure_type'
    ]
    header_len = len(headers)

    data = []
    for obj in fetch_data(api_key, url):
        headers = data_utils.append_missing_headers(obj, headers)
        data.append(obj)
    headers = data_utils.order_headers(headers, header_len)
    headers = data_utils.normalize_headers(headers)

    sheet = wb.create_sheet('relationships')
    sheet.append([header[0].split('.')[-1] for header in headers])

    for obj in data:
        sheet.append([data_utils.get_attr(obj, attr[1]) for attr in headers])
